import os
import io  # <-- هذا هو السطر الذي تحتاجه بشدة لإصلاح الخطأ
import openpyxl
from django.conf import settings
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import Eleve
# 1. الصفحة الرئيسية: البحث أولاً ثم الأوائل الثلاثة
def home_view(request):
    series = ['SN', 'M', 'LO', 'LM', 'TM', 'TS', 'LA']
    top_students = {}
    for s in series:
        top_students[s] = Eleve.objects.filter(serie=s, statut__icontains='admis').order_by('-moyenne')[:3]

    student_result = None
    rank_in_serie = None
    searched = False

    if 'q' in request.GET:
        searched = True
        query = request.GET.get('q', '').strip()
        if query:
            try:
                student_result = Eleve.objects.get(num_table=query)
                status_val = str(student_result.statut or '').lower()
                
                # حساب الترتيب فقط إذا كان ناجحاً
                if 'admis' in status_val or 'ناجح' in status_val:
                    higher_count = Eleve.objects.filter(
                        serie=student_result.serie,
                        statut__icontains='admis',
                        moyenne__gt=student_result.moyenne
                    ).count()
                    rank_in_serie = higher_count + 1
            except Eleve.DoesNotExist:
                student_result = None

    context = {
        'top_students': top_students,
        'student': student_result,
        'rank': rank_in_serie,
        'searched': searched  # هذا المتغير سيتحكم في إخفاء كروت المتفوقين
    }
    return render(request, 'home.html', context)


# 2. صفحة الفرز والإحصائيات الذكية والمتجاوبة
# 2. صفحة الفرز والإحصائيات الذكية والمتجاوبة (نسخة الفلترة الديناميكية تلقائياً)
def stats_view(request):
    # 1. استقبال قيم الفلاتر من الطلب (GET)
    selected_wilaya = request.GET.get('wilaya', '')
    selected_serie = request.GET.get('serie', '')
    selected_school = request.GET.get('etablissement', '')
    selected_status = request.GET.get('status', '')

    # 2. بناء قوائم الخيارات بشكل ديناميكي من قاعدة البيانات لكافة الفلاتر
    wilayas = Eleve.objects.values_list('wilaya', flat=True).distinct().order_by('wilaya')
    series_list = Eleve.objects.values_list('serie', flat=True).distinct().order_by('serie')
    status_list = Eleve.objects.values_list('statut', flat=True).distinct().order_by('statut')

    # 3. التصفية الذكية للمؤسسات حسب الولاية والشعبة لتسهيل الاختيار التلقائي
    school_filter = Eleve.objects.all()
    if selected_wilaya:
        school_filter = school_filter.filter(wilaya=selected_wilaya)
    if selected_serie:
        school_filter = school_filter.filter(serie=selected_serie)
        
    schools = school_filter.values_list('etablissement', flat=True).distinct().order_by('etablissement')

    # 4. فلترة مشروطة: لا يتم جلب أي بيانات (None) إلا بعد اختيار الولاية كشرط للبدء
    students = Eleve.objects.none()
    has_filtered = False

    # تفعيل البحث والنتائج فقط في حال قام المستخدم باختيار ولاية محددة
    if selected_wilaya:
        has_filtered = True
        students = Eleve.objects.filter(wilaya=selected_wilaya)
        
        if selected_serie:
            students = students.filter(serie=selected_serie)
        if selected_school:
            students = students.filter(etablissement=selected_school)
            
        if selected_status:
            if selected_status in ['Admis', 'ناجح']:
                students = students.filter(statut__icontains='admis')
            elif selected_status in ['Ajourné', 'راسب']:
                students = students.filter(statut__icontains='ajourn')
            elif selected_status in ['Sessionaire', 'دورة ثانية', 'دورة تكميلية']:
                students = students.filter(statut__icontains='sess')
            elif selected_status in ['Absent', 'غائب']:
                students = students.filter(statut__icontains='abs')
            else:
                students = students.filter(statut=selected_status)

    total_count = students.count() if has_filtered else 0

    context = {
        'wilayas': wilayas,
        'series_list': series_list,
        'schools': schools,
        'status_list': status_list,
        'students': students,
        'total_count': total_count,
        'has_filtered': has_filtered,
        'selected_wilaya': selected_wilaya,
        'selected_serie': selected_serie,
        'selected_school': selected_school,
        'selected_status': selected_status,
    }
    return render(request, 'stats.html', context)

# 3. رفع ملف الإكسيل وقراءة العناوين
# 1. الدالة الأولى: ترفع الملف وتقرأ العناوين (وتمرر الملف مشفراً داخل الـ Session دون حفظه كملف على القرص)
def upload_excel_view(request):
    if request.method == "POST" and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        
        try:
            # قراءة الملف مباشرة من الذاكرة (RAM) لضمان السرعة وعدم الاعتماد على القرص الصلب المؤقت
            file_in_memory = excel_file.read()
            workbook = openpyxl.load_workbook(io.BytesIO(file_in_memory), read_only=True)
            sheet = workbook.active
            headers = [str(cell.value).strip() for cell in next(sheet.iter_rows(max_row=1))]
            workbook.close()
            
            # حل مشكلة السيرفر: نحفظ محتوى الملف بشكل آمن داخل "جلسة المستخدم" Session ليمر بأمان للصفحة التالية
            request.session['uploaded_excel_data'] = file_in_memory.decode('latin-1') 
            
            return render(request, 'mapping.html', {'headers': headers})
        except Exception as e:
            messages.error(request, f"خطأ في قراءة ملف الإكسيل: {e}")
            return render(request, 'upload.html')
            
    return render(request, 'upload.html')


# 2. الدالة الثانية: معالجة واستيراد البيانات مع حماية الذاكرة العشوائية للسيرفر من الـ Crash
def import_mapped_data_view(request):
    if request.method == "POST":
        # جلب بيانات الملف من الـ Session بأمان وضمان عدم اختفائها
        file_data_raw = request.session.get('uploaded_excel_data')
        if not file_data_raw:
            messages.error(request, "انتهت صلاحية الجلسة أو لم يتم العثور على الملف، يرجى إعادة الرفع.")
            return redirect('upload_excel')
            
        try:
            # جلب ترتيب الأعمدة التي اختارها المستخدم
            idx_table = int(request.POST.get('col_table'))
            idx_nom = int(request.POST.get('col_nom'))
            idx_wilaya = int(request.POST.get('col_wilaya'))
            idx_etablissement = int(request.POST.get('col_etablissement'))
            idx_centre = int(request.POST.get('col_centre'))
            idx_serie = int(request.POST.get('col_serie'))
            idx_moyenne = int(request.POST.get('col_moyenne'))
            idx_statut = int(request.POST.get('col_statut'))
        except:
            messages.error(request, "تأكد من اختيار كافة الأعمدة.")
            return redirect('upload_excel')

        try:
            # إعادة تحويل البيانات النصية إلى بايتس لقراءتها بواسطة openpyxl مجدداً
            file_bytes = io.BytesIO(file_data_raw.encode('latin-1'))
            workbook = openpyxl.load_workbook(file_bytes, read_only=True, data_only=True)
            sheet = workbook.active
            
            # مسح البيانات القديمة فوراً لتفريغ مساحة الداتابيز للبيانات الجديدة
            Eleve.objects.all().delete()
            
            students_batch = []
            total_saved = 0
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row: continue
                required_len = max(idx_table, idx_nom, idx_wilaya, idx_etablissement, idx_centre, idx_serie, idx_moyenne, idx_statut) + 1
                if len(row) < required_len: continue
                
                num_table = str(row[idx_table]).strip()
                if not num_table or num_table.lower() == "none": continue
                
                moyenne_raw = str(row[idx_moyenne]).replace(',', '.').strip()
                try: moyenne_val = float(moyenne_raw)
                except: moyenne_val = 0.0
                
                # إعداد الحالة (منطقك الذكي والمطور لمعالجة كافة الحالات)
                raw_statut = str(row[idx_statut]).strip().lower()
                if 'adm' in raw_statut or 'ناجح' in raw_statut:
                    final_statut = 'Admis'
                elif 'sess' in raw_statut or 'ثاني' in raw_statut or 'تكميل' in raw_statut:
                    final_statut = 'Sessionaire'
                elif 'abs' in raw_statut or 'غائب' in raw_statut or 'غايب' in raw_statut:
                    final_statut = 'Absent'
                elif 'ajou' in raw_statut or 'راسب' in raw_statut:
                    final_statut = 'Ajourné'
                else:
                    if moyenne_val >= 10.0: final_statut = 'Admis'
                    else: final_statut = 'Ajourné'

                # صمام أمان المعدل
                if moyenne_val >= 10.0 and final_statut == 'Ajourné':
                    final_statut = 'Admis'

                student = Eleve(
                    num_table=num_table,
                    nom_complet=str(row[idx_nom]).strip(),
                    wilaya=str(row[idx_wilaya]).strip(),
                    etablissement=str(row[idx_etablissement]).strip(),
                    centre=str(row[idx_centre]).strip(),
                    serie=str(row[idx_serie]).strip().upper(),
                    moyenne=moyenne_val,
                    statut=final_statut
                )
                students_batch.append(student)
                
                # الاستراتيجية الأهم: كلما تجمعت دفعة من 500 طالب، يتم حفظهم فوراً في قاعدة البيانات وتفريغ الذاكرة
                if len(students_batch) >= 500:
                    Eleve.objects.bulk_create(students_batch)
                    total_saved += len(students_batch)
                    students_batch = [] # مسح وتصفير القائمة فوراً لتفريغ الـ RAM

            # حفظ المتبقي الأخير من الطلاب إن وُجد (أقل من 500)
            if students_batch:
                Eleve.objects.bulk_create(students_batch)
                total_saved += len(students_batch)

            workbook.close()
            
            # تنظيف الـ Session وحذف بيانات ملف الإكسيل منها بعد انتهاء الحفظ بنجاح
            if 'uploaded_excel_data' in request.session:
                del request.session['uploaded_excel_data']
                
            messages.success(request, f"تمت المزامنة بنجاح وحفظ {total_saved} طالب بكافة حالاتهم المختلفة بكفاءة عالية.")
            
        except Exception as e:
            messages.error(request, f"حدث خطأ أثناء استيراد البيانات: {e}")

        return redirect('upload_excel')
    return redirect('upload_excel')